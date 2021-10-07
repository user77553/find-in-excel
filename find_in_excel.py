import openpyxl as xl

try:
    file = open("temp", "r")
    filename = file.read()
    file.close()
except:
    filename = input("Enter a filename of excel file (list.xlsx): ")
    file = open("temp", "w")
    file.write(filename)
    file.close()

search_string_input = input("Enter a search string: ")


def word_finder(search_string, ws):
    result = []
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            try:
                v = ws.cell(i, j).value.lower()
                if v.find(search_string.lower()) != -1:
                    result.append(ws.cell(i, j))
            except:
                pass
    return result


if filename.find("xlsx") != -1 and isinstance(search_string_input, str):
    print(f"File: {filename}")
    try:
        wb = xl.load_workbook(filename)
        search_str = []
        not_found = True
        counter = 0

        for sheet in wb.worksheets:
            search_str = word_finder(search_string_input, sheet)
            for unit in search_str:
                if hasattr(unit, 'value'):
                    print(f"Found: '{unit.value}' in --> {sheet.title}.")
                    not_found = False
                    counter += 1
                else:
                    pass

        if not_found:
            print(f"'{search_string_input}' wasn't found in file {filename}.")
        else:
            print(f"Total {counter}.")
    except:
        print("File not found")
else:
    print("Invalid filename, try again.")